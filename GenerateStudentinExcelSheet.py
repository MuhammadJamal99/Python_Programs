import openpyxl
import random

# Load the Excel workbook
workbook = openpyxl.load_workbook('Adding_Multiple_Students_Template.xlsx')

# Select the sheet where you want to insert records
sheet = workbook['Sheet1']

# Define the number of records to insert
num_records = int(input("Enter the number of Wanted records:"))

# Ask the user to enter the value for the portal variable
portal = input("Enter the value for the portal: ")

#portal = '@test123'

# Ask the user to enter the value for the portal variable
students_groups_input = input("Enter the values for student groups (separated by commas): ")

# Parse the input into a list
students_groups_array = [int(x.strip()) for x in students_groups_input.split(',')]

#students_groups_array = [72347, 7234747]

current_row = 2;
print(f"Started Current Row: {current_row}");
# Insert random values into the columns
for i in range(num_records):
    first_name = 'Student';
    last_name = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz', k=5));
    email = f'{first_name}.{last_name}{portal}'
    group_code = random.choice(students_groups_array)
    student_code = f'{first_name}{i}'
    
    sheet.cell(row=current_row, column=1, value=first_name)
    sheet.cell(row=current_row, column=2, value=last_name)
    sheet.cell(row=current_row, column=3, value=email)
    sheet.cell(row=current_row, column=4, value=group_code)
    sheet.cell(row=current_row, column=5, value=student_code)
    print(f"Current Row: {current_row}");
    current_row += 1

# Save the changes to the workbook
workbook.save('Adding_Multiple_Students_Template_Updated.xlsx');